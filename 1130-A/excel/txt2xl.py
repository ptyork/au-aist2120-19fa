import openpyxl as xl
import os

# create a new workbook
wb = xl.Workbook()
#wb.remove(wb.active)
wb.remove(wb.worksheets[0])

# for each file
for filename in os.listdir():
    if not filename.endswith('.txt'):
        continue
    if not os.path.isfile(filename):
        continue

#   get assignment name
    assignment = filename[:-4]

#   create a worksheet (and give it a name)
    ws = wb.create_sheet(assignment)
    '''
            ws["A1"].value = 10
            ws.cell(row=1, column=1).value = 10
            ws.cell(row=1, column=1, value=10)
            for cell in ws["A1":"B3"]:
                print cell.value
            for row in ws.rows:
                for cell in row:
                    print cell.value
                    cell.value = 10
    '''

#   for each line
    with open(filename) as file:
        rownum = 1
        for line in file:
            # alan.alda,75\n
            line = line.strip()
            # alan.alda,75
            
#     get id and grade
            parts = line.split(',')
            # ['alan.alda', '75']
            if len(parts) != 2:
                continue   # to the next line in the file
            
            id = parts[0]
            try:
                grade = float(parts[1])
            except:
                #continue
                grade = 0

#     add a row with id and grade to the worksheet
            ws.cell(row=rownum, column=1, value=id)
            ws.cell(row=rownum, column=2, value=grade)
            rownum += 1

# save the workbook
wb.save('grades.xlsx')