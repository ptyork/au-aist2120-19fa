import openpyxl as xl
import os

# create a new excel workbook
wb = xl.Workbook()
#wb.remove(wb.active)
#del wb.worksheets[0]
wb.remove(wb.worksheets[0])

# for every .txt file in the folder
for filename in os.listdir():
    if not filename.endswith('.txt'):
        continue
    if not os.path.isfile(filename):
        continue
#    get name of assignment
    assignment = filename[:-4]

#    create a worksheet
    ws = wb.create_sheet(assignment)

    '''
    ws["A1"].value = 123
    for row in ws.rows:
        for col in row:
            col.value = 5
    ws.cell(row=1, column=1).value = 7
    ws.cell(row=1, column=1, value=9)
    '''

#    for each line in the file
    with open(filename) as file:
        rownum = 1
        for line in file:
            line = line.strip()
            parts = line.split(',')

            if len(parts) != 2:
                continue

            id = parts[0]
            try:
                grade = float(parts[1])
            except:
                #continue
                grade = 0.0
            
#           create a row with that data
            ws.cell(row=rownum, column=1, value=id)
            ws.cell(row=rownum, column=2, value=grade)

            rownum += 1

wb.save('grades.xlsx')
